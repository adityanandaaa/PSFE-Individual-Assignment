import pandas as pd
import json
import logging
from datetime import datetime
from functools import lru_cache
from finance_app.config import CURRENCIES_FILE
from finance_app.logging_config import get_logger
from finance_app.models import FinancialRecord
from pydantic import ValidationError

# Get module logger
logger = get_logger(__name__)

@lru_cache(maxsize=1)
def load_currencies():
    """Load currencies from JSON file (cached)."""
    try:
        # Attempt to open and parse the currencies JSON file
        with open(CURRENCIES_FILE, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        # Log error if file does not exist and return empty list as fallback
        logger.error("currencies.json not found.")
        return []

def is_valid_income(income):
    """Validate income input - must be numeric and positive."""
    try:
        # Convert input to float (handles both numeric and string inputs)
        val = float(income)
        # Check if income is greater than zero (negative or zero income is invalid)
        return val > 0
    except (ValueError, TypeError):
        # Return False if conversion fails (non-numeric input)
        return False

def get_currency_symbol(currencies, code):
    """Get currency symbol by code from the currencies list."""
    # Find currency object with matching code (case-sensitive)
    currency = next((c for c in currencies if c['code'] == code), None)
    # Return symbol if found, otherwise None
    return currency['symbol'] if currency else None

def validate_file(file_path):
    """Validate the uploaded Excel file structure and content.
    
    Returns:
        tuple: (is_valid: bool, result: pd.DataFrame or list of (row_num, error_msg) tuples)
    """
    # Check file size limit (50 MB) to prevent memory issues with very large uploads
    MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB in bytes
    try:
        # Get file size - handle both file paths and BytesIO objects
        if hasattr(file_path, 'getbuffer'):
            # BytesIO object from Streamlit file uploader
            file_size = len(file_path.getbuffer())
        else:
            # File path string
            import os
            file_size = os.path.getsize(file_path)
        
        if file_size > MAX_FILE_SIZE:
            size_mb = file_size / 1024 / 1024
            logger.warning(f"Uploaded file too large: {size_mb:.1f} MB (limit: 50 MB)")
            return False, [(0, f"File too large ({size_mb:.1f} MB > 50 MB limit)")]
    except Exception as e:
        # Log that we couldn't check size, but don't stop as validation will check content
        logger.debug(f"Could not determine file size before validation: {str(e)}")
    
    try:
        # Load Excel file using openpyxl engine for better compatibility
        # Use context manager to ensure file is properly closed after reading
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Read data into DataFrame from worksheet
        data = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = row
            else:
                data.append(row)
        
        df = pd.DataFrame(data, columns=headers)

        # Normalize: keep only required columns, drop fully empty rows
        required_cols = ['Date', 'Name', 'Type', 'Amount', 'Category']
        # Remove rows where all required fields are empty/NaN/whitespace
        df = (
            df[required_cols]
            .replace(r'^\s*$', pd.NA, regex=True)
            .dropna(how='all')
            .reset_index(drop=True)
        )
        wb.close()  # Explicitly close the workbook to release file lock
        
        errors = []  # List of (row_num, error_msg) tuples
        
        # Check if all required columns exist
        if not all(col in df.columns for col in required_cols):
            errors.append((0, "Missing required columns."))
            return False, errors
        
        # Iterate through each row to validate data using Pydantic
        for idx, row in df.iterrows():
            try:
                # Convert row to dictionary for Pydantic validation
                row_dict = row.to_dict()
                
                # Pre-processing for Pydantic (Amount string cleanup)
                if isinstance(row_dict.get('Amount'), str):
                    row_dict['Amount'] = row_dict['Amount'].replace(',', '')

                # Validate using Pydantic model
                FinancialRecord.model_validate(row_dict)
                
            except ValidationError as e:
                # Extract the first error message for the row
                for error in e.errors():
                    # Format field-specific errors into human-readable messages
                    field = str(error['loc'][0]).lower()
                    # Map Pydantic errors back to the app's error message style
                    if field == 'name':
                        if 'max_length' in error['type']:
                            errors.append((idx+1, "Name too long (max 150 chars)."))
                        else:
                            errors.append((idx+1, "Invalid name."))
                    elif field == 'type':
                        errors.append((idx+1, "Invalid type."))
                    elif field == 'amount':
                        errors.append((idx+1, "Invalid amount."))
                    elif field == 'date':
                        errors.append((idx+1, "Invalid date format."))
                    elif field == 'category':
                        errors.append((idx+1, "Invalid category."))
                    else:
                        errors.append((idx+1, f"Invalid {field}"))
                    break # Only report the first error per row for clarity
            except Exception as e:
                logger.error(f"Unexpected error validating row {idx+1}: {str(e)}")
                errors.append((idx+1, "Unexpected error in data format."))
        
        # Return results based on whether any errors were found
        if errors:
            return False, errors
        return True, df
    except Exception as e:
        # Handle any file reading errors
        return False, [(0, f"Error reading file: {str(e)}")]

def analyze_data(df, income):
    """Analyze financial data and calculate budget allocation.
    
    Aggregates spending by Type (Needs, Wants, Savings) and calculates
    the top 5 spending categories within Wants for detailed analysis.
    
    Args:
        df: DataFrame with columns [Date, Name, Type, Amount, Category]
        income: Total monthly income for reference
        
    Returns:
        tuple: (needs, wants, savings, top_wants_series)
    """
    # Safety check for non-DataFrame input (contract enforcement)
    if not isinstance(df, pd.DataFrame):
        logger.error(f"analyze_data received non-DataFrame input: {type(df)}")
        # Return zeros and empty series as safe fallback to prevent app crash
        return 0, 0, 0, pd.Series(dtype=float)

    # Normalize category names to lowercase for consistent grouping
    df['Category'] = df['Category'].str.lower()
    
    # Group by Type and sum amounts for each budget category
    buckets = df.groupby('Type')['Amount'].sum()
    
    # Extract total amounts for each budget category (use .get() for safety if missing)
    needs = buckets.get('Needs', 0)
    wants = buckets.get('Wants', 0)
    savings = buckets.get('Savings', 0)
    
    # === TOP WANTS ANALYSIS ===
    # Filter for only 'Wants' type transactions
    wants_df = df[df['Type'] == 'Wants']
    # Group wants by category and get top 5 by total amount
    top_wants = wants_df.groupby('Category')['Amount'].sum().nlargest(5)
    
    # Return aggregated budget data
    return needs, wants, savings, top_wants

def calculate_health_score(income, needs, wants, savings):
    """Calculate monthly budget health score based on 50/30/20 budgeting model.
    
    Uses directional deviation logic to penalize only risky behaviors:
    - Needs overspend (> 50%) - small penalty
    - Wants overspend (> 30%) - heavier penalty
    - Under-saving (< 20%) - heaviest penalty
    
    Weighted deviations:
    Dn = max(0, n - 0.5)      # Needs overspend
    Dw = max(0, w - 0.3)      # Wants overspend
    Ds = max(0, 0.2 - s)      # Under-saving
    Score = 100 × (1 - (0.2×Dn + 0.5×Dw + 0.6×Ds))
    
    Hard risk rules:
    - If savings < 0: Score = 0 (financial danger - debt/overspend)
    - If needs > 75%: Score -= 10 (extreme overspend)
    
    Final score is clamped to [0, 100].
    
    Args:
        income: Total monthly income
        needs: Total needs spending
        wants: Total wants spending
        savings: Total savings
        
    Returns:
        int: Health score from 0 to 100
    """
    # === STEP 1: CALCULATE RATIOS ===
    # Prevent division by zero
    if income <= 0:
        return 0
    
    # Convert spending to ratios (as fractions of income)
    n = needs / income      # Needs ratio
    w = wants / income      # Wants ratio
    s = savings / income    # Savings ratio
    
    # === STEP 2: DIRECTIONAL DEVIATIONS (only penalize risk) ===
    # Only measure overspend/under-save, not underspend/over-save
    Dn = max(0, n - 0.5)      # Needs overspend (penalty only if > 50%)
    Dw = max(0, w - 0.3)      # Wants overspend (penalty only if > 30%)
    Ds = max(0, 0.2 - s)      # Under-saving (penalty only if < 20%)
    
    # === STEP 3: WEIGHTED SCORE CALCULATION ===
    # Weights: Needs 0.2, Wants 0.5, Savings 0.6
    # Higher weights on Wants and Savings reflect their importance to financial health
    score = 100 * (1 - (0.2 * Dn + 0.5 * Dw + 0.6 * Ds))
    
    # === STEP 4: HARD MONTHLY RISK RULES ===
    # Severe penalty: negative savings means debt or overspending
    if savings < 0:
        score = 0
    
    # Penalize extreme needs overspend (> 75%)
    if n > 0.75:
        score -= 10
    
    # === STEP 5: CLAMP TO 0–100 ===
    score = max(0, min(100, score))
    
    return int(score)