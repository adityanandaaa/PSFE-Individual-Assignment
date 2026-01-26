import pandas as pd
import json
import logging
from datetime import datetime
from modules.config import CURRENCIES_FILE

def load_currencies():
    """Load currencies from JSON file."""
    try:
        # Attempt to open and parse the currencies JSON file
        with open(CURRENCIES_FILE, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        # Log error if file does not exist and return empty list as fallback
        logging.error("currencies.json not found.")
        return []

def is_valid_income(income):
    """Validate income input - must be numeric and positive."""
    try:
        # Convert input to float (handles both numeric and string inputs)
        val = float(income)
        # Check if income is greater than zero (negative or zero income is invalid)
        return val > 0
    except (ValueError, TypeError):
        # Return False if conversion fails (non-numeric input)
        return False

def get_currency_symbol(currencies, code):
    """Get currency symbol by code from the currencies list."""
    # Find currency object with matching code (case-sensitive)
    currency = next((c for c in currencies if c['code'] == code), None)
    # Return symbol if found, otherwise None
    return currency['symbol'] if currency else None

def validate_file(file_path):
    """Validate the uploaded Excel file structure and content.
    
    Returns:
        tuple: (is_valid: bool, result: pd.DataFrame or list of error strings)
    """
    try:
        # Load Excel file using openpyxl engine for better compatibility
        # Use context manager to ensure file is properly closed after reading
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Read data into DataFrame from worksheet
        data = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = row
            else:
                data.append(row)
        
        df = pd.DataFrame(data, columns=headers)
        wb.close()  # Explicitly close the workbook to release file lock
        
        errors = []
        
        # Define required column names for the financial tracker
        required_cols = ['Date', 'Name', 'Type', 'Amount', 'Category']
        
        # Check if all required columns exist
        if not all(col in df.columns for col in required_cols):
            errors.append("Missing required columns.")
        
        # Iterate through each row to validate data
        for idx, row in df.iterrows():
            # === DATE VALIDATION ===
            if pd.isna(row['Date']):
                errors.append(f"Row {idx+1}: Date is empty.")
            else:
                # Validate date format (expects dd/mm/yyyy)
                try:
                    datetime.strptime(str(row['Date']), '%d/%m/%Y')
                except ValueError:
                    errors.append(f"Row {idx+1}: Invalid date format.")
            
            # === NAME VALIDATION ===
            # Check: not empty, is string, max 150 chars, contains alphanumeric + spaces
            if pd.isna(row['Name']) or not isinstance(row['Name'], str) or len(str(row['Name'])) > 150 or str(row['Name']).replace(' ', '').isalnum() == False:
                errors.append(f"Row {idx+1}: Invalid name.")
            
            # === TYPE VALIDATION ===
            # Type must be one of the three budget categories
            if row['Type'] not in ['Needs', 'Wants', 'Savings']:
                errors.append(f"Row {idx+1}: Invalid type.")
            
            # === AMOUNT VALIDATION ===
            try:
                # Convert amount to float for numeric validation
                amt = float(row['Amount'])
                # Amount must be positive
                if amt <= 0:
                    errors.append(f"Row {idx+1}: Invalid amount.")
                else:
                    # Check decimal places - maximum 3 decimal places allowed (e.g., 10.999)
                    s = str(row['Amount'])
                    if '.' in s:
                        dec = s.split('.')[-1]
                        if len(dec) > 3:
                            errors.append(f"Row {idx+1}: Invalid amount.")
            except Exception:
                # Amount conversion failed
                errors.append(f"Row {idx+1}: Invalid amount.")
            
            # === CATEGORY VALIDATION ===
            # Check: not empty, is string, max 150 chars
            if pd.isna(row['Category']) or not isinstance(row['Category'], str) or len(str(row['Category'])) > 150:
                errors.append(f"Row {idx+1}: Invalid category.")
        
        # Return results based on whether any errors were found
        if errors:
            return False, errors
        return True, df
    except Exception as e:
        # Handle any file reading errors
        return False, [f"Error reading file: {str(e)}"]

def analyze_data(df, income):
    """Analyze financial data and calculate budget allocation.
    
    Aggregates spending by Type (Needs, Wants, Savings) and calculates
    the top 5 spending categories within Wants for detailed analysis.
    
    Args:
        df: DataFrame with columns [Date, Name, Type, Amount, Category]
        income: Total monthly income for reference
        
    Returns:
        tuple: (needs, wants, savings, top_wants_series)
    """
    # Normalize category names to lowercase for consistent grouping
    df['Category'] = df['Category'].str.lower()
    
    # Group by Type and sum amounts for each budget category
    buckets = df.groupby('Type')['Amount'].sum()
    
    # Extract total amounts for each budget category (use .get() for safety if missing)
    needs = buckets.get('Needs', 0)
    wants = buckets.get('Wants', 0)
    savings = buckets.get('Savings', 0)
    
    # === TOP WANTS ANALYSIS ===
    # Filter for only 'Wants' type transactions
    wants_df = df[df['Type'] == 'Wants']
    # Group wants by category and get top 5 by total amount
    top_wants = wants_df.groupby('Category')['Amount'].sum().nlargest(5)
    
    # Return aggregated budget data
    return needs, wants, savings, top_wants

def calculate_health_score(income, needs, wants, savings):
    """Calculate financial health score based on 50/30/20 budgeting model.
    
    Uses mathematical formula to evaluate how well spending aligns with targets:
    - Needs: 50% target (weight: 0.35)
    - Wants: 30% target (weight: 0.15)
    - Savings: 20% target (weight: 0.50)
    
    Formula:
    Score = 100 × (1 - [0.35|n-0.5| + 0.15|w-0.3| + 0.50|s-0.2|])
    
    Adjustment rules applied after core calculation:
    - If savings < 5%: Score -= 20
    - If needs > 70%: Score -= 15
    - If savings < 0: Score = 0
    
    Final score is clamped to [0, 100].
    
    Args:
        income: Total monthly income
        needs: Total needs spending
        wants: Total wants spending
        savings: Total savings
        
    Returns:
        int: Health score from 0 to 100
    """
    # === CALCULATE RATIOS ===
    # Prevent division by zero
    if income <= 0:
        return 0
    
    # Convert spending to ratios (as fractions of income)
    n = needs / income  # Needs ratio
    w = wants / income  # Wants ratio
    s = savings / income  # Savings ratio
    
    # === CORE SCORE CALCULATION ===
    # Calculate weighted deviation from targets
    # Needs target: 0.5, Wants target: 0.3, Savings target: 0.2
    deviation = 0.35 * abs(n - 0.5) + 0.15 * abs(w - 0.3) + 0.50 * abs(s - 0.2)
    
    # Convert deviation to score (0-100)
    score = 100 * (1 - deviation)
    
    # === ADJUSTMENT RULES ===
    # Penalize insufficient savings (< 5%)
    if s < 0.05:
        score -= 20
    
    # Penalize excessive needs spending (> 70%)
    if n > 0.70:
        score -= 15
    
    # Severe penalty: negative savings means debt or overspending
    if savings < 0:
        score = 0
    
    # === FINAL CONSTRAINT ===
    # Clamp score to valid range [0, 100]
    score = max(0, min(100, score))
    
    return int(score)